from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from database import get_db
from models import DeliveryPartners, MasterCustomers, DeliveryLogs, Payments
from auth.auth_utils import require_admin
from datetime import date
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import io


router = APIRouter()


@router.get("/export/dsr/month/{month}")
def export_dsr(month: str, db: Session = Depends(get_db), current_user: dict = Depends(require_admin)):

    year, mon = int(month.split("-")[0]), int(month.split("-")[1])
    if mon == 12:
        next_month = date(year + 1, 1, 1)
    else:
        next_month = date(year, mon + 1, 1)
    days_in_month = (date(year, mon + 1, 1) - date(year, mon, 1)).days if mon != 12 else 31

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    # styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    total_font = Font(bold=True)
    center = Alignment(horizontal="center")

    partners = db.query(DeliveryPartners).all()

    for partner in partners:
        customers = db.query(MasterCustomers).filter(
            MasterCustomers.delivery_partner_id == partner.id,
            MasterCustomers.status == "active"
        ).all()

        if not customers:
            continue

        # get all logs for this partner this month
        logs = db.query(DeliveryLogs).filter(
            DeliveryLogs.delivery_partner_id == partner.id,
            func.to_char(DeliveryLogs.delivery_date, "YYYY-MM") == month,
            DeliveryLogs.delivery_status == "done"  # only count done deliveries
        ).all()

        # build lookup {customer_id: {day: litres}}
        log_lookup = {}
        for log in logs:
            cid = log.customer_id
            day = log.delivery_date.day
            if cid not in log_lookup:
                log_lookup[cid] = {}
            log_lookup[cid][day] = log.quantity_delivered

        # create sheet — sheet name = partner name (max 31 chars for Excel)
        ws = wb.create_sheet(title=partner.name[:31])

        # header row
        headers = ["Customer Name", "Address"] + [str(d) for d in range(1, days_in_month + 1)] + ["Total Litres"]
        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center

        # data rows
        for row_idx, customer in enumerate(customers, start=2):
            ws.cell(row=row_idx, column=1, value=customer.name)
            ws.cell(row=row_idx, column=2, value=customer.address)

            total = 0.0
            for day in range(1, days_in_month + 1):
                litres = log_lookup.get(customer.id, {}).get(day, 0)
                ws.cell(row=row_idx, column=2 + day, value=litres if litres else "")
                total += litres if litres else 0

            # total litres column
            total_cell = ws.cell(row=row_idx, column=2 + days_in_month + 1, value=total)
            total_cell.font = total_font

        # column widths
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 20
        for col in range(3, days_in_month + 4):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 5

    # save to memory buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=DSR_{month}.xlsx"}
    )


@router.get("/export/billing/month/{month}")
def export_billing(month: str, db: Session = Depends(get_db), current_user: dict = Depends(require_admin)):

    year, mon = int(month.split("-")[0]), int(month.split("-")[1])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Billing {month}"

    # styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="375623", end_color="375623", fill_type="solid")
    total_font = Font(bold=True)
    center = Alignment(horizontal="center")
    red_font = Font(bold=True, color="FF0000")

    # header row
    headers = [
        "Sr No.", "Customer Name", "Address", "Phone",
        "Milk Type", "Rate", "Total Litres",
        "Previous Balance", "Total Amount", "Grand Total",
        "Total Paid", "Balance"
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    # get all active customers
    customers = db.query(MasterCustomers).filter(
        MasterCustomers.status == "active"
    ).all()

    for idx, customer in enumerate(customers, start=1):

        # total litres delivered this month (done only)
        total_litres = db.query(func.sum(DeliveryLogs.quantity_delivered)).filter(
            DeliveryLogs.customer_id == customer.id,
            func.to_char(DeliveryLogs.delivery_date, "YYYY-MM") == month,
            DeliveryLogs.delivery_status == "done"
        ).scalar() or 0.0

        # bill calculation
        total_amount = total_litres * customer.rate
        previous_balance = customer.previous_balance or 0.0
        grand_total = total_amount + previous_balance

        # total paid this month
        total_paid = db.query(func.sum(Payments.amount)).filter(
            Payments.customer_id == customer.id,
            Payments.for_month == month
        ).scalar() or 0.0

        balance = grand_total - total_paid

        row = idx + 1
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=customer.name)
        ws.cell(row=row, column=3, value=customer.address)
        ws.cell(row=row, column=4, value=customer.phone)
        ws.cell(row=row, column=5, value=customer.milk_type)
        ws.cell(row=row, column=6, value=customer.rate)
        ws.cell(row=row, column=7, value=round(total_litres, 2))
        ws.cell(row=row, column=8, value=round(previous_balance, 2))
        ws.cell(row=row, column=9, value=round(total_amount, 2))
        ws.cell(row=row, column=10, value=round(grand_total, 2))
        ws.cell(row=row, column=11, value=round(total_paid, 2))

        # balance cell — red if still owes money
        balance_cell = ws.cell(row=row, column=12, value=round(balance, 2))
        if balance > 0:
            balance_cell.font = red_font
        else:
            balance_cell.font = total_font

    # column widths
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 8
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 16
    ws.column_dimensions["I"].width = 14
    ws.column_dimensions["J"].width = 12
    ws.column_dimensions["K"].width = 12
    ws.column_dimensions["L"].width = 12

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=Billing_{month}.xlsx"}
    )